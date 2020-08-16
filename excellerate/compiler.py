from nmigen import *
from nmigen.sim.pysim import *
from openpyxl import load_workbook
from collections import defaultdict

from fixedpoint import Q, QArray
import parser
from functions import Location, Cell, Sum

class CellDict(defaultdict):
    def __init__(self, nint, nfrac, signed):
        super().__init__()
        self.nint = nint
        self.nfrac = nfrac
        self.signed = signed

    def __missing__(self, key):
        name = f"{key.sheet}_{key.col}_{key.row}"
        sig = Signal(Shape(self.nint+self.nfrac, self.signed), name=name)
        ready = Signal(name=name)
        cell = Cell(Q(self.nint, self.nfrac, self.signed, sig), ready)
        self[key] = cell
        return cell



class Spreadsheet(Elaboratable):

    def __init__(self, workbook, nint=16, nfrac=16, signed=True):
        self.nint = nint
        self.nfrac = nfrac
        self.signed = signed
        self.workbook = workbook
        self.submodules = []

        self.cells = CellDict(nint, nfrac, signed)

    def elaborate(self, platform):
        m = Module()
        for sheet in wb.sheetnames:
            for row in wb[sheet]:
                for cell in row:
                    loc = Location(sheet, cell.column, cell.row)
                    ast = parser.parse(str(cell.value))
                    sig = self.compile_cell(loc, ast)
                    if sig is not None:
                        cell = self.cells[Location(sheet, cell.column, cell.row)]
                        m.d.sync += cell.value.eq(sig.value)
                        m.d.sync += cell.ready.eq(sig.ready)
        m.submodules += self.submodules
        return m

    def compile_cell(self, cell, ast):
        if ast is None:
            return None
        elif isinstance(ast, tuple):
            op, left, right = ast
            lcell = self.compile_cell(cell, left)
            rcell = self.compile_cell(cell, right)
            if op == '+':
                res = lcell.value + rcell.value
            elif op == '-':
                res = lcell.value - rcell.value
            elif op == '*':
                res = lcell.value * rcell.value
            elif op == '/':
                res = lcell.value / rcell.value
            elif op == '^':
                res = lcell.value ** rcell.value
            elif op == '>':
                res = lcell.value > rcell.value
            elif op == '>=':
                res = lcell.value >= rcell.value
            elif op == '<':
                res = lcell.value < rcell.value
            elif op == '<=':
                res = lcell.value <= rcell.value
            elif op == '=':
                res = lcell.value == rcell.value
            elif op == '<>':
                res = lcell.value != rcell.value
            else:
                raise NameError(f"operator {op} not handled")
            ready = lcell.ready | rcell.ready
            return Cell(res, ready)
        elif isinstance(ast, float):
            return Cell(
                Q.from_float(ast, self.nint, self.nfrac, self.signed),
                Const(0)
            )
        elif isinstance(ast, parser.Array):
            return Cell(
                QArray([
                    QArray([
                        Q.from_float(i, self.nint, self.nfrac, self.signed)
                        for i in row])
                    for row in ast.elements]),
                Const(0)
            )
        elif isinstance(ast, parser.Range):
            sheet = ast.sheet or cell.sheet
            min_col, min_row, max_col, max_row = ast.boundaries
            if min_col==max_col and min_row==max_row:
                return self.cells[Location(sheet, min_col, min_row)]
            return Cell(
                QArray([
                    QArray([
                        self.cells[Location(sheet, col, row)].value
                        for col in range(min_col, max_col+1)])
                    for row in range(min_row, max_row+1)]),
                Cat(self.cells[Location(sheet, col, row)].ready
                    for col in range(min_col, max_col+1)
                    for row in range(min_row, max_row+1)).any()
            )
        elif isinstance(ast, parser.Function):
            if ast.name == "SUM":
                args = [self.compile_cell(cell, arg) for arg in ast.args]
                summer = Sum(*args)
                self.submodules.append(summer)
                return summer.result
        else:
            raise TypeError(f"{ast} is not of a supported type")

if __name__ == '__main__':
    wb = load_workbook('simple.xlsx')
    spr = Spreadsheet(wb)
    sim = Simulator(spr)
    def testbench():
        for i in range(50):
            yield Tick()
            print("###### TICK ######")
            for loc, cell in spr.cells.items():
                res = yield cell.value.signal
                print(loc, cell.value.to_float(res))

    sim.add_clock(1e-6)
    sim.add_process(testbench)
    with sim.write_vcd("test.vcd", "test.gtkw", traces=[c.value.signal for c in spr.cells.values()]):
        sim.run()
